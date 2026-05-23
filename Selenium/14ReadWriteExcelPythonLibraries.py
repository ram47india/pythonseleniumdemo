import openpyxl
dic = {}
book = openpyxl.load_workbook("C:\\Users\\SESA644934\\Downloads\\TestData.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Ram"
print("Before Saving Excel:",sheet.cell(row=2, column=2).value)
# book.save("C:\\Users\\SESA644934\\Downloads\\TestData.xlsx")
print("After Saving Excel:",sheet.cell(row=2, column=2).value)
print(sheet.max_row, sheet.max_column)

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value, end=" | ")
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row = i, column = 1).value == "testcase2":
        for j in range(2,sheet.max_column+1):
            dic[sheet.cell(row=1,column=j).value] = sheet.cell(row=i, column=j).value
print(dic)
