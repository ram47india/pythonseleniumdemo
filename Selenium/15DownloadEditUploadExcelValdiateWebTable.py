import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions

file_path = "C:\\Users\\SESA644934\\Downloads\\download.xlsx"
fruit_name = "Apple"
new_price = 400
col_name = "price"

def update_excel(filepath, fruitname, colname, newprice):
    import openpyxl
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    dic = {}
    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1,column=i).value == colname:
            dic["col"] = i
    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i,column=j).value == fruitname:
                dic["row"] = i
    sheet.cell(row=dic["row"],column=dic["col"]).value=newprice
    workbook.save(filepath)

driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.maximize_window()
# Download the excel file
driver.find_element(By.ID,"downloadButton").click()
time.sleep(3)  #Wait for file to download

# Edit the downloaded excel file
update_excel(file_path, fruit_name, col_name, new_price)

# Upload the edited excel file
driver.find_element(By.XPATH,"//input[@id='fileinput']").send_keys(file_path)
locator = (By.XPATH,"//div[text()='Updated Excel Data Successfully.']")
wait = WebDriverWait(driver,5)
wait.until(expected_conditions.visibility_of_element_located(locator))
print(driver.find_element(*locator).text)

# Validate the web table data
# validate the price of Apple is updated to 400(same way based on any columm value like input "Spring" validate the price/color/fruit name)
input_column_value = "Apple"
output_expected_column = "Price"
output_column_id = driver.find_element(By.XPATH,"//div[text()='"+output_expected_column+"']").get_attribute("data-column-id")
# print(output_column_id)
# actual_Output = driver.find_element(By.XPATH,"//div[text()='Apple']/parent::div/parent::div/div[@id='cell-4-undefined']").text
actual_output = driver.find_element(By.XPATH,"//div[text()='"+input_column_value+"']/parent::div/parent::div/div[@id='cell-"+output_column_id+"-undefined']").text
print(actual_output)
assert actual_output == "400"